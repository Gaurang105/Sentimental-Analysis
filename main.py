import os
import nltk
import re
import openpyxl
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords

nltk.download('punkt')
nltk.download('stopwords')

# Load Input.xlsx file
input_wb = openpyxl.load_workbook('Input.xlsx')
input_ws = input_wb.active

# Load or create Output.xlsx file
output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.append(['URL_ID', 'URL', 'POSITIVE SCORE', 'NEGATIVE SCORE', 'POLARITY SCORE', 'SUBJECTIVITY SCORE', 
                  'AVG SENTENCE LENGTH', 'PERCENTAGE OF COMPLEX WORDS', 'FOG INDEX', 'WORD COUNT', 
                  'COMPLEX WORD COUNT', 'SYLLABLE PER WORD', 'PERSONAL PRONOUNS', 'AVG WORD LENGTH'])

# Folder with articles and positive-negative words
articles_dir = 'scraped-articles'
pos_neg_words_dir = 'positive-negative words'

# Load positive and negative words
with open(os.path.join(pos_neg_words_dir, 'positive-words.txt'), 'r', encoding='utf-8') as f:
    positive_words = set(f.read().splitlines())
with open(os.path.join(pos_neg_words_dir, 'negative-words.txt'), 'r', encoding='utf-8') as f:
    negative_words = set(f.read().splitlines())

# Define stopwords
stop_words = set(stopwords.words('english'))

# Define function to count syllables in a word
def syllable_count(word):
    word = word.lower()
    count = 0
    vowels = "aeiouy"
    if word[0] in vowels:
        count += 1
    for index in range(1, len(word)):
        if word[index] in vowels and word[index-1] not in vowels:
            count += 1
    if word.endswith("e"):
        count -= 1
    if word.endswith("le"):
        count += 1
    if count == 0:
        count += 1
    return count

# Process each row in the input worksheet
for row in input_ws.iter_rows(min_row=2, values_only=True):
    url_id, url = row
    filename = f"{int(url_id)}.txt"
    filepath = os.path.join(articles_dir, filename)

    if not os.path.exists(filepath):
        print(f'File {filepath} not found, skipping.')
        continue

    print(f'Trying to open: {filepath}')
    with open(filepath, 'r', encoding='utf-8') as f:
        text = f.read()

    # Tokenize text
    sentences = sent_tokenize(text)
    tokens = word_tokenize(text)

    # Calculate scores
    positive_score = sum(word in positive_words for word in tokens)
    negative_score = sum(word in negative_words for word in tokens)

    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    subjectivity_score = (positive_score + negative_score) / ((len(tokens)) + 0.000001)

    # Readability calculations
    average_sentence_length = len(tokens) / len(sentences)
    complex_words_count = sum(1 for word in tokens if syllable_count(word) > 2)
    percentage_complex_words = complex_words_count / len(tokens)
    fog_index = 0.4 * (average_sentence_length + percentage_complex_words)
    
    # Word count after removing stop words and punctuation
    cleaned_words = [word for word in tokens if word.isalnum() and word not in stop_words]
    word_count = len(cleaned_words)

    # Syllables per word
    syllables_per_word = sum(syllable_count(word) for word in cleaned_words) / word_count
    
    # Personal pronouns
    personal_pronouns_count = len(re.findall(r'\b(I|we|my|ours|us)\b', text, re.IGNORECASE))

    # Average word length
    average_word_length = sum(len(word) for word in cleaned_words) / word_count

    # Append results to output worksheet
    output_ws.append([url_id, url, positive_score, negative_score, polarity_score, subjectivity_score, 
                      average_sentence_length, percentage_complex_words, fog_index, word_count, 
                      complex_words_count, syllables_per_word, personal_pronouns_count, average_word_length])

# Save output workbook
output_wb.save('Output Data Structure.xlsx')

